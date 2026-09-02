import asyncio
import random
import re
import logging
import os
import subprocess
import webbrowser
from datetime import datetime, timedelta
import openpyxl
from tkinter import filedialog, messagebox
import tkinter as tk
from playwright.async_api import async_playwright, TimeoutError as PWTimeout
try:
    from playwright_stealth import Stealth
    STEALTH_AVAILABLE = True
except ImportError:
    STEALTH_AVAILABLE = False

__version__ = "3.10.2"
VERSION_URL = "https://raw.githubusercontent.com/vipywk-lab/DH-checker/main/bx_checker.py"
NAS_PATH    = r"\\10.223.120.38\종합통제\24. 승무계획팀\29.자동화\DH 조회 자동화"
GITHUB_URL  = "https://github.com/vipywk-lab/DH-checker"

# 실행 시 콘솔에 표시되는 이번 버전 변경사항 (유저용 — 기술 용어 지양, 짧게)
LATEST_CHANGELOG = (
    "  - 대한항공 조회 시 뜨는 쿠키 확인 창을 자동으로 넘기도록 수정"
)

# 클라우드플레어 감지 키워드 (전역 — 모든 항공사 조회 함수에서 공유)
CF_KEYWORDS = ["보안 확인 수행 중", "사람인지 확인하십시오", "Checking your browser",
               "DDoS protection", "보안 서비스", "악의적인 봇", "Cloudflare"]


def _is_reliable_result(flt_found, route_found):
    """
    편명·구간을 둘 다 못 읽었으면 실제 예약 정보를 읽은 게 아니라
    (알려지지 않은 오류 문구의 실패 페이지 등) 실패일 가능성이 높음.
    실패 키워드 목록에 없다고 무조건 "확인완료"로 넘기지 않기 위한 안전장치.
    """
    return not (flt_found == "편명미확인" and route_found == "구간미확인")


# 조회 성공 시 결과 페이지 텍스트를 보관 — 같은 PNR의 다른 탑승객이
# 실제로 그 예약에 포함돼 있는지 확인하는 데 사용
_page_cache = {}


def _remember_page(pnr, airline, page_text):
    """조회 성공한 결과 페이지 텍스트를 보관 (같은 PNR 동승자 검증용)"""
    try:
        _page_cache[(pnr, airline)] = str(page_text)
    except Exception:
        pass


def _name_in_page(target, page_text):
    """
    해당 탑승객이 결과 페이지에 실제로 있는지 확인.
    한글명 / 영문명(성·이름 각각) 어느 쪽이든 발견되면 True.
    ※ 사이트가 이름을 가리는 경우(홍*동 등)엔 못 찾을 수 있으며,
      그때는 캐시를 쓰지 않고 개별 조회하므로 안전하게 동작함.
    """
    if not page_text:
        return False
    norm = re.sub(r'\s+', '', str(page_text)).upper()

    # 한글명 (동명이인 구분자 A/B 제거한 형태)
    kor = re.sub(r'[A-Za-z]+$', '', str(target.get("kor_name", ""))).strip()
    if kor and re.sub(r'\s+', '', kor) in norm:
        return True

    # 영문명 "YU/DONGYUN" → 성·이름 모두 페이지에 있어야 인정
    eng = str(target.get("eng_name", "")).strip().upper()
    if eng:
        parts = [p for p in re.split(r'[/\s]+', eng) if p]
        if parts and all(re.sub(r'\s+', '', p) in norm for p in parts):
            return True

    return False
# ==========================================
# 체인지로그
# v3.10.2 (2026-09-02) — 대한항공 쿠키 확인창 자동 처리
#   - 대한항공 조회 페이지에서 뜨는 "모든 쿠키 허용" 확인창을 자동으로 닫도록 수정
# v3.10.1 (2026-08-26) — [중요] 캐시 오판정 수정 (사용자 리포트)
#   - 같은 예약번호(PNR)면 이전 결과를 그대로 재사용하면서 그 사람이 실제로
#     해당 예약 명단에 있는지는 확인하지 않아, 엉뚱한 사람도 '확인완료'로
#     넘어가던 문제 수정
#   - 이제 캐시 사용 전 결과 화면에 해당 탑승객(한글명/영문명)이 있는지 확인하고,
#     확인되지 않으면 캐시를 쓰지 않고 개별 조회함 (화면 표시: [캐시·동승확인])
# v3.10.0 (2026-08-26) — 이어서 조회 기능 추가
#   - 실행 시 이미 '확인완료'된 건이 있으면 건너뛸지 묻는 팝업 표시
#     (한 달치 400건 규모는 1시간 가까이 걸려, 중단 시 처음부터 다시 하던 문제 해소)
#   - 오류·불일치·PNR오류·미조회 건은 건너뛰지 않고 항상 다시 조회
#   - 건너뛴 건의 기존 결과는 그대로 유지되고 요약 시트에도 정상 반영됨
# v3.9.4 (2026-08-26) — 파일 선택 관련 안내 개선
#   - 조회 시작 시 실제로 읽는 엑셀 파일의 전체 경로를 콘솔에 표시
#     (화면에서 편집한 파일과 다른 파일을 선택해 "대상 0건"이 뜨던 혼란 방지)
#   - 대상 0건일 때 원인 체크리스트(파일 확인/저장 여부/조회 범위) 안내 추가
# v3.9.3 (2026-07-24) — 안정성 보강 (전체 재검토)
#   - 이름 칸이 비었거나 알파벳만 있는 경우 IndexError로 그 건이 오류나던 문제 방어
#   - 조회 결과가 엑셀 행과 매칭 안 되면(이름/PNR 불일치) 조용히 빈칸으로 남던 것을
#     콘솔에 매칭 실패 건을 안내하도록 개선
#   - GitHub 파일에서 버전 정보를 못 찾을 때 안내 출력 (버전체크 무력화 방지)
# v3.9.2 (2026-07-24) — 전체 재검토로 발견한 버그 수정
#   - 조회 도중 예외가 나면 프로그램이 통째로 죽고 그때까지의 결과가
#     하나도 저장되지 않던 문제 수정 (이제 항상 엑셀 저장 + 미조회 건 표시)
#   - 제주항공 재조회(영문/타임아웃) 분기가 빠져 있어 프로그램이 멈추던 문제 수정
#   - 대상 0건일 때 오류로 종료되던 문제 수정
#   - PNR오류가 오류 건수에 중복 집계되던 문제 수정
# v3.9.1 (2026-07-24) — [매우 중요] PNR 오판정 버그 수정
#   - 에어부산/대한항공: 조회마다 같은 브라우저 탭을 재사용하고 있었는데,
#     잘못된 PNR 조회가 실패해도 화면이 안 바뀌면 직전 사람의 결과 화면이
#     그대로 남아있어 그걸 "확인완료"로 잘못 읽는 문제가 있었음
#     → 조회마다 새 탭 사용 + 조회한 PNR이 실제 결과 화면에 있는지 확인하는
#       이중 안전장치 추가 (파라타항공에도 PNR 검증 추가)
# v3.9.0 (2026-07-07) — [중요] 오판정 방지 안전장치 추가
#   - 에어부산/대한항공/진에어/파라타항공/제주항공 5개 항공사 전부 해당:
#     실패 문구를 못 찾았다고 해서 무조건 "확인완료"로 반환하지 않도록 수정.
#     편명·구간을 둘 다 못 읽은 경우(=제대로 된 예약 페이지를 읽은 게 아닐 가능성)
#     "PNR오류"로 표시해 즉시 확인하도록 변경
# v3.8.x (2026-07-04) — 제주항공 조회 기능 추가
#   - PNR/성명 자동입력, 조회결과 자동판정(날짜·구간까지 정확히 확인)
#   - 달력 날짜 선택만 사람이 클릭(안내 팝업 표시), 나머지는 자동 진행
#   - 사이트 광고 팝업 자동 닫기, 구버전 안내에 NAS 경로/GitHub 링크 반영
# v3.7.0 (2026-07-04) — [중요] 진에어 날짜 오판정 수정
#   - 예약일을 출발일로 착각해서 멀쩡한 예약도 "날짜불일치"로 잘못 뜨던 문제 수정
# v3.6.0 (2026-07-03) — 제주항공 조회 최초 추가
# v3.5.0~3.5.2 (2026-07-03)
#   - GitHub 버전체크 연동(구버전이면 실행 시 자동 안내 후 차단)
#   - 실행창 제목표시줄에 버전 자동 표시
# v3.4.0~3.4.1 (2026-07-03) — 티웨이항공 조회 추가
#   - 자동조회 불가 사이트라 팝업 안내 + 항목별 [복사] 버튼으로 지원
# v3.3.0~3.3.2 (2026-07-01~03)
#   - 진에어 조회 오류 수정(정상 예약도 PNR오류로 뜨던 문제)
#   - 조회 범위 선택 1/2/3 버튼 복원, 오류 로그 저장 위치 고정
# v3.2.1 이하 — 초기 안정화 (생략)
# ==========================================
# Playwright Chromium 최초 1회 자동 설치
# ==========================================
def check_for_update():
    """GitHub raw URL에서 최신 버전 확인 — 구버전이면 실행 차단"""
    import urllib.request
    try:
        with urllib.request.urlopen(VERSION_URL, timeout=5) as resp:
            found_version = False
            for line in resp.read().decode("utf-8").splitlines():
                if line.startswith("__version__"):
                    found_version = True
                    latest = line.split("=")[1].strip().strip('"').strip("'")

                    def _ver(v):
                        try:
                            return tuple(int(x) for x in v.split("."))
                        except Exception:
                            return (0,)

                    if _ver(latest) > _ver(__version__):
                        print(f"\n{'!'*50}")
                        print(f"  ⚠️  업데이트 필요: 현재 v{__version__} → 최신 v{latest}")
                        print(f"  최신 파일로 교체가 필요합니다. 아래 둘 중 하나로 받아주세요.")
                        print(f"  1) NAS: {NAS_PATH}")
                        print(f"  2) GitHub: {GITHUB_URL}")
                        print(f"{'!'*50}\n")
                        input("업데이트 후 다시 실행해주세요. 엔터 누르면 종료...")
                        raise SystemExit("구버전 실행 차단")
                    else:
                        print(f"✅ 최신 버전입니다 (v{__version__})")
                    return
            # 파일은 받았으나 __version__ 줄을 못 찾은 경우 (포맷 변경 등)
            if not found_version:
                print("⚠️  버전 정보를 확인하지 못했습니다 (버전 체크 건너뜀)\n")
    except SystemExit:
        raise
    except Exception:
        print("⚠️  버전 확인 실패 (네트워크 연결 없음 — 무시하고 계속 진행)\n")


check_for_update()

chromium_path = os.path.expanduser("~\\AppData\\Local\\ms-playwright")
if not os.path.exists(chromium_path):
    print("기반 시스템(브라우저)을 설치 중입니다. 최초 1회만 진행되며 시간이 조금 걸릴 수 있습니다...")
    try:
        subprocess.run(["playwright", "install", "chromium"], check=True)
        print("설치 완료!\n")
    except Exception as e:
        print(f"설치 중 오류가 발생했습니다: {e}")

# ==========================================
# 로깅 설정 (오류 발생 시 텍스트 파일로 저장)
# ==========================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_PATH = os.path.join(SCRIPT_DIR, f"에러로그_{datetime.now().strftime('%Y%m%d')}.txt")
logging.basicConfig(
    filename=LOG_PATH,
    level=logging.ERROR,
    format='%(asctime)s [%(levelname)s] %(message)s',
    encoding='utf-8'
)

root = tk.Tk()
root.withdraw()
messagebox.showinfo("안내", "xlsm 파일을 선택해주세요")
EXCEL_PATH = filedialog.askopenfilename(
    title="xlsm 파일 선택",
    filetypes=[("Excel files", "*.xlsm *.xlsx")]
)
if not EXCEL_PATH:
    raise SystemExit("파일을 선택하지 않았습니다.")

SHEET_NAME = "검증대상"
RESULT_COL = 8
DETAIL_COL = 9
BX_URL     = "https://www.airbusan.com/web/individual/reserve/index"
KE_URL     = "https://www.koreanair.com/reservation/search"
LJ_URL     = "https://www.jinair.com/booking/index"
WE_URL     = "https://www.parataair.com/ko/login/viewLogin.do?tab=2#"
TW_URL     = "https://www.twayair.com/app/reservation/searchMemberBooking"
JJ_URL     = "https://www.jejuair.net/ko/ibe/mypage/viewOnOffReservationList.do"

# 국내/국제선 공항코드 (전역 — 여러 항공사 조회 함수에서 공유)
AIRPORT_CODES = (
    r'PUS|GMP|ICN|CJU|TAE|CJJ|HIN|RSU|KPO|MWX'
    r'|CNX|BKK|HKT|NRT|HND|KIX|NGO|CTS|FUK|OKA'
    r'|DAD|SGN|HAN|CXR|PQC|CEB|KLO|TAG|MNL'
    r'|TPE|HKG|MFM|SIN|DPS|GUM'
)
HEADLESS   = False
DELAY_MIN  = 1.0
DELAY_MAX  = 2.0

DOMESTIC_AIRPORTS = {"PUS","CJU","TAE","CJJ","HIN","RSU","KPO","MWX","GMP","ICN"}


def ask_resume(done_count, total_count):
    """
    이미 '확인완료'된 건이 있으면 건너뛸지 묻는 팝업.
    반환값: True(건너뛰기) / False(전체 다시 조회)
    """
    result = [None]
    popup = tk.Toplevel(root)
    popup.title("이어서 조회")
    popup.resizable(False, False)
    popup.attributes("-topmost", True)
    popup.grab_set()

    tk.Label(
        popup,
        text=(
            f"이미 조회가 끝난 건이 {done_count}건 있습니다.\n"
            f"(전체 {total_count}건 중)\n\n"
            "건너뛰고 나머지만 조회하면 시간을 크게 줄일 수 있습니다.\n"
            "※ 오류·불일치·미조회 건은 건너뛰지 않고 다시 조회합니다.\n\n"
            "어떻게 할까요?"
        ),
        justify="left", padx=20, pady=15
    ).pack()

    btn_frame = tk.Frame(popup)
    btn_frame.pack(pady=(0, 15))

    tk.Button(
        btn_frame, text=f"  이어서 조회 ({total_count - done_count}건)  ", width=22,
        command=lambda: [result.__setitem__(0, True), popup.destroy()]
    ).pack(side="left", padx=6)
    tk.Button(
        btn_frame, text=f"  처음부터 다시 ({total_count}건)  ", width=22,
        command=lambda: [result.__setitem__(0, False), popup.destroy()]
    ).pack(side="left", padx=6)

    popup.wait_window()
    # 팝업을 그냥 닫으면 안전하게 '이어서 조회'로 처리
    return True if result[0] is None else result[0]


def get_check_mode():
    """실행 시 조회 범위 선택 팝업 — 1:5일 / 2:이번달말 / 3:다음달"""
    import calendar
    today = datetime.now()

    # 이번달 말
    this_last = calendar.monthrange(today.year, today.month)[1]
    days_to_eom = this_last - today.day

    # 다음달
    if today.month == 12:
        next_year, next_month = today.year + 1, 1
    else:
        next_year, next_month = today.year, today.month + 1
    next_last = calendar.monthrange(next_year, next_month)[1]

    result = [None]
    popup = tk.Toplevel(root)
    popup.title("조회 범위 선택")
    popup.resizable(False, False)
    popup.grab_set()

    tk.Label(
        popup,
        text=(
            "조회 범위를 선택하세요.\n\n"
            f"  1.  오늘부터 5일\n"
            f"  2.  이번달 말까지 ({today.month}월 {this_last}일, 약 {days_to_eom}일)\n"
            f"  3.  다음달 ({next_year}년 {next_month}월 1일 ~ {next_last}일)\n\n"
            "※ 2·3번은 딜레이가 자동으로 늘어납니다."
        ),
        justify="left",
        padx=20, pady=15
    ).pack()

    btn_frame = tk.Frame(popup)
    btn_frame.pack(pady=(0, 15))

    for n in (1, 2, 3):
        tk.Button(
            btn_frame, text=f"  {n}번  ", width=8,
            command=lambda v=n: [result.__setitem__(0, v), popup.destroy()]
        ).pack(side="left", padx=8)

    popup.wait_window()

    if result[0] == 2:
        end = datetime(today.year, today.month, this_last)
        return "this", today, end, 3.0, 6.0
    elif result[0] == 3:
        start = datetime(next_year, next_month, 1)
        end   = datetime(next_year, next_month, next_last)
        return "next", start, end, 3.0, 6.0
    else:  # 1번 또는 팝업 강제 종료
        end = today + timedelta(days=5)
        return "5d", today, end, 1.0, 2.0


def split_korean_name(name):
    name = name.strip()
    # 동명이인 구분자 제거: 이경수A → 이경수, 박지연B → 박지연
    stripped = re.sub(r'[A-Za-z]+$', '', name).strip()
    # 접미사 제거 후 비었으면(알파벳만 있거나 빈 값) 원본 사용, 그래도 비면 빈 문자열 반환
    base = stripped if stripped else name
    if not base:
        return "", ""
    return base[0], base[1:]


def parse_dep_date(dep_time_str):
    s = str(dep_time_str).strip()
    try:
        return datetime.strptime(s[:7], "%d%b%y")
    except:
        pass
    try:
        return datetime.strptime(s[:9], "%d-%b-%y")
    except:
        pass
    try:
        n = float(s)
        if 40000 < n < 60000:
            return datetime(1899, 12, 30) + timedelta(days=int(n))
    except:
        pass
    return None


def is_within_check_range(dep_time_str, start_date, end_date):
    dep_date = parse_dep_date(str(dep_time_str))
    if not dep_date:
        return True
    return start_date.replace(hour=0, minute=0, second=0, microsecond=0) <= dep_date <= end_date


def is_international(dep, arr):
    return dep not in DOMESTIC_AIRPORTS or arr not in DOMESTIC_AIRPORTS


def load_targets(path, sheet, start_date, end_date):
    wb = openpyxl.load_workbook(path, keep_vba=True)
    if sheet not in wb.sheetnames:
        messagebox.showerror(
            "시트 없음",
            f"선택한 파일에 [{sheet}] 시트가 없습니다.\n\n"
            f"① DH_자동화.xlsm에서 VBA 매크로를 먼저 실행해 [{sheet}] 시트를 생성하세요.\n"
            f"② 올바른 파일을 선택했는지 확인하세요.\n\n"
            f"선택한 파일: {os.path.basename(path)}\n"
            f"현재 시트 목록: {', '.join(wb.sheetnames)}"
        )
        raise SystemExit(f"[{sheet}] 시트 없음 → 프로그램 종료")
    ws = wb[sheet]
    targets = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = (list(row) + [None]*9)[:9]
        kor_name, airline, pnr, dep, arr, dep_time, eng_name = vals[:7]
        prev_result, prev_detail = vals[7], vals[8]
        if not all([kor_name, airline, pnr]):
            continue
        if airline not in ("에어부산", "대한항공", "진에어", "제주항공", "파라타항공", "티웨이항공"):
            continue
        if not re.match(r'^[A-Z0-9]{6}$', str(pnr).strip().upper()):
            continue
        if not is_within_check_range(str(dep_time or ""), start_date, end_date):
            continue
        last, first = split_korean_name(str(kor_name))
        targets.append({
            "kor_name": str(kor_name),
            "airline" : str(airline),
            "last"    : last,
            "first"   : first,
            "eng_name": str(eng_name).strip().replace("-", " ") if eng_name else "",
            "pnr"     : str(pnr).strip().upper(),
            "dep"     : str(dep or ""),
            "arr"     : str(arr or ""),
            "dep_time": str(dep_time or ""),
            "result"  : None,
            "detail"  : None,
            "prev_result": str(prev_result) if prev_result else "",
            "prev_detail": str(prev_detail) if prev_detail else "",
        })
    return targets


def save_results(path, sheet, targets):
    """PNR 기반으로 정확하게 매칭해서 저장 + 확인필요 요약 시트 생성"""
    wb = openpyxl.load_workbook(path, keep_vba=True)
    ws = wb[sheet]
    ws.cell(1, RESULT_COL).value = "검증결과"
    ws.cell(1, DETAIL_COL).value = "조회내용"

    # PNR + 항공사 + 이름 조합으로 정확히 매칭
    result_map = {}
    for t in targets:
        key = (t["pnr"], t["airline"], t["kor_name"])
        result_map[key] = (t["result"], t["detail"])

    matched_keys = set()
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        airline  = row[1].value
        pnr      = str(row[2].value).strip().upper() if row[2].value else ""
        kor_name = str(row[0].value).strip() if row[0].value else ""
        if airline not in ("에어부산", "대한항공", "진에어", "제주항공", "파라타항공", "티웨이항공"):
            continue
        key = (pnr, airline, kor_name)
        if key in result_map:
            result, detail = result_map[key]
            ws.cell(row_idx, RESULT_COL).value = result
            ws.cell(row_idx, DETAIL_COL).value = detail
            matched_keys.add(key)

    # 조회는 했으나 엑셀 행과 매칭 안 된 건 경고 (이름/PNR이 조회 중 미묘하게 달라진 경우)
    unmatched = [t for t in targets if (t["pnr"], t["airline"], t["kor_name"]) not in matched_keys]
    if unmatched:
        print(f"\n⚠️  결과 {len(unmatched)}건이 엑셀 행과 매칭되지 않아 기록 못 함:")
        for t in unmatched:
            print(f"     - {t['kor_name']} | {t['airline']} | {t['pnr']}")
        print("   (엑셀 원본의 이름/PNR과 조회 데이터가 다른지 확인 필요)")

    # ── 확인필요 요약 시트 생성 ──
    summary_name = "확인필요_요약"
    if summary_name in wb.sheetnames:
        del wb[summary_name]

    ws_summary = wb.create_sheet(summary_name)
    ws_summary.append(["탑승객", "항공사", "PNR", "검증결과", "조회내용"])
    ws_summary.row_dimensions[1].height = 20

    error_count = 0
    for t in targets:
        res = str(t["result"])
        if any(kw in res for kw in ["불일치", "예약없음", "PNR오류", "오류", "타임아웃", "수동확인필요", "미조회"]):
            ws_summary.append([
                t["kor_name"],
                t["airline"],
                t["pnr"],
                t["result"],
                t["detail"]
            ])
            error_count += 1

    if error_count == 0:
        ws_summary.append(["✅ 모든 예약이 정상적으로 확인되었습니다!"])
    else:
        # 헤더 볼드
        from openpyxl.styles import Font
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)

    ws_summary.column_dimensions["A"].width = 12
    ws_summary.column_dimensions["B"].width = 12
    ws_summary.column_dimensions["C"].width = 10
    ws_summary.column_dimensions["D"].width = 15
    ws_summary.column_dimensions["E"].width = 40

    # ── 저장 (엑셀 열려있으면 팝업 안내) ──
    while True:
        try:
            wb.save(path)
            print(f"\n저장 완료: {path}")
            print(f"→ [확인필요_요약] 시트에서 {error_count}건 확인하세요!" if error_count > 0 else "→ 모든 예약 정상!")
            break
        except Exception:
            messagebox.showerror(
                "저장 오류",
                f"엑셀 파일이 열려있거나 저장할 수 없습니다!\n\n{path}\n\n파일을 닫고 '확인'을 눌러주세요."
            )


async def check_bx(page, target):
    pnr      = target["pnr"]
    eng_name = target.get("eng_name", "")
    dep      = target["dep"]
    arr      = target["arr"]

    intl = is_international(dep, arr)

    if intl and eng_name:
        parts = eng_name.split("/")
        last  = parts[0].strip() if len(parts) >= 1 else target["last"]
        first = parts[1].strip() if len(parts) >= 2 else target["first"]
    else:
        last  = target["last"]
        first = target["first"]

    # 조회마다 새 탭 사용 — 이전 사람의 결과 화면이 남아있는 상태에서
    # 이번 조회가 실패하면 그 잔류 화면을 잘못 읽어버리는 문제 방지
    bx_page = await page.context.new_page()
    try:
        await bx_page.goto(BX_URL, wait_until="domcontentloaded", timeout=20000)
        await bx_page.wait_for_timeout(1500)

        # 클라우드플레어 감지 → 사람이 직접 캡챠 풀도록 안내
        body_check = await bx_page.inner_text("body")
        if any(kw in body_check for kw in CF_KEYWORDS):
            print(f"\n{'='*50}")
            print(f"  ⚠️  [에어부산] 클라우드플레어 보안 확인이 필요합니다!")
            print(f"  → 열린 브라우저에서 '사람인지 확인하십시오' 체크박스를 클릭해주세요.")
            print(f"  → 완료 후 여기서 엔터를 눌러주세요.")
            print(f"{'='*50}")
            await asyncio.get_event_loop().run_in_executor(None, input, "  [확인 후 엔터] ")
            # 통과됐는지 재확인
            body_check2 = await bx_page.inner_text("body")
            if any(kw in body_check2 for kw in CF_KEYWORDS):
                return "⏱️ 타임아웃", "클라우드플레어 차단 미해제 → 재실행 필요"

        await bx_page.click("text=예약번호로 조회", timeout=5000)
        await bx_page.wait_for_timeout(800)

        await bx_page.locator("input[placeholder*='예약번호']").first.fill(pnr)
        await bx_page.wait_for_timeout(300)
        await bx_page.locator("input[placeholder='성']").first.fill(last)
        await bx_page.wait_for_timeout(300)
        await bx_page.locator("input[placeholder='이름']").first.fill(first)
        await bx_page.wait_for_timeout(300)

        await bx_page.evaluate("document.querySelector('.buttonOfflineCheckin').click()")
        await bx_page.wait_for_timeout(2000)

        body_text = await bx_page.inner_text("body")
        if "해당 예약번호가 확인되지 않습니다" in body_text:
            try:
                await bx_page.click("button:has-text('확인')", timeout=2000)
            except:
                pass
            return "❌ PNR오류", "해당 예약번호 확인 불가"

        try:
            await bx_page.wait_for_selector("text=항공권 구매완료", timeout=10000)
        except:
            pass
        await bx_page.wait_for_timeout(1000)

        html_content = await bx_page.inner_text("body")

        # 결과 파싱 전 클라우드플레어 재체크
        if any(kw in html_content for kw in CF_KEYWORDS):
            return "⏱️ 타임아웃", "클라우드플레어 차단 → 재실행 필요"
        if any(kw in html_content for kw in ["조회 결과가 없", "예약 내역이 없", "일치하는 예약"]):
            return "❌ 예약없음", "조회결과 없음"

        # 조회한 PNR이 실제로 결과 화면에 있는지 확인 (잔류 화면 오판정 방지 2중 안전장치)
        if pnr.upper() not in html_content.upper():
            return "❌ PNR오류", "조회한 PNR이 결과 화면에서 확인되지 않음"

        flt_match   = re.search(r'BX\s*\d{3,4}', html_content)
        date_match  = re.search(r'(\d{4}-\d{2}-\d{2})', html_content)

        # 구간: 한글 도시명 (국내 + 에어부산 국제선 취항지)
        BX_CITIES = (
            '부산|서울|김포|제주|대구|광주|청주|인천'
            '|후쿠오카|오사카|삿포로|도쿄|나리타'
            '|다낭|나트랑|냐짱|보홀|세부|칼리보|비엔티안|치앙마이'
            '|타이베이|가오슝|홍콩|마카오|칭다오|옌지|장자제|시안|상하이'
        )
        route_match = re.search(
            rf'({BX_CITIES}).{{1,15}}({BX_CITIES})',
            html_content
        )

        if route_match:
            route_found = route_match.group()
        else:
            # fallback: 공항코드로 매칭 (한글 도시명 미표시 대비)
            codes = re.findall(
                r'(?<![A-Z0-9])(PUS|GMP|ICN|CJU|TAE|CJJ|HIN|RSU|KPO|MWX'
                r'|FUK|KIX|CTS|NRT|HND|NGO'
                r'|DAD|CXR|TAG|CEB|KLO|VTE|CNX'
                r'|TPE|KHH|HKG|MFM|TAO|YNJ|DYG|XIY|PVG)(?![A-Z0-9])',
                html_content
            )
            route_found = f"{codes[0]}→{codes[1]}" if len(codes) >= 2 else "구간미확인"

        flt_found   = flt_match.group().replace(" ", "") if flt_match else "편명미확인"
        date_found  = date_match.group() if date_match else "날짜미확인"

        detail = f"{flt_found} | {date_found} | {route_found}"

        dep_date = parse_dep_date(target["dep_time"])
        mismatch = []
        if dep_date and date_found != "날짜미확인":
            try:
                site_date = datetime.strptime(date_found, "%Y-%m-%d")
                if dep_date.date() != site_date.date():
                    mismatch.append(
                        f"날짜불일치(PDC:{dep_date.strftime('%m/%d')} vs 사이트:{site_date.strftime('%m/%d')})"
                    )
            except:
                pass

        if mismatch:
            return "⚠️ 불일치", detail + " | " + " / ".join(mismatch)

        if not _is_reliable_result(flt_found, route_found):
            return "❌ PNR오류", f"예약 확인 불가 (편명/구간 모두 미확인) | {detail}"

        # 같은 PNR 동승자 검증용으로 결과 페이지 보관
        _remember_page(pnr, target["airline"], html_content)
        return "✅ 확인완료", detail

    except PWTimeout:
        return "⏱️ 타임아웃", "재시도 필요"
    except Exception as e:
        logging.error(f"에어부산 조회 실패 | PNR: {pnr} | 탑승객: {last}{first}", exc_info=True)
        return "💥 오류", "시스템 로그 확인 필요"
    finally:
        await bx_page.close()


async def check_ke(page, target):
    pnr      = target["pnr"]
    eng_name = target.get("eng_name", "")
    dep      = target["dep"]
    arr      = target["arr"]

    # 국내선=한글, 국제선=영문 (BX/LJ와 동일)
    intl = is_international(dep, arr)
    if intl and eng_name:
        parts = eng_name.split("/")
        last  = parts[0].strip() if len(parts) >= 1 else target["last"]
        first = parts[1].strip() if len(parts) >= 2 else target["first"]
    else:
        last  = target["last"]
        first = target["first"]

    dep_date = parse_dep_date(target["dep_time"])
    if not dep_date:
        return "💥 오류", "출발일 파싱 실패"

    ke_page = await page.context.new_page()
    try:
        await ke_page.goto(KE_URL, wait_until="domcontentloaded", timeout=20000)
        await ke_page.wait_for_timeout(2000)

        try:
            await ke_page.click("button:has-text('동의합니다')", timeout=3000)
            await ke_page.wait_for_timeout(500)
        except:
            pass

        try:
            await ke_page.click("button.-confirm:has-text('모든 쿠키 허용')", timeout=3000)
            await ke_page.wait_for_timeout(500)
        except:
            pass

        await ke_page.locator("input[maxlength='13']").first.fill(pnr)
        await ke_page.wait_for_timeout(300)

        await ke_page.click("button[data-dialog-id='#dialog-datepicker1']", timeout=5000)
        await ke_page.wait_for_timeout(1500)

        dep_day   = str(dep_date.day)
        dep_month = dep_date.month
        dep_year  = dep_date.year

        await ke_page.evaluate(f"""
            (function() {{
                var tds = document.querySelectorAll('td.datepicker__td.-available[role="button"]');
                for (var td of tds) {{
                    var span = td.querySelector('span[aria-hidden="true"]');
                    if (!span) continue;
                    if (span.textContent.trim() !== '{dep_day}') continue;
                    var container = td.closest('[id*="month{dep_year}{dep_month:02d}"]');
                    if (container) {{
                        td.click();
                        return;
                    }}
                }}
                for (var td of tds) {{
                    var span = td.querySelector('span[aria-hidden="true"]');
                    if (span && span.textContent.trim() === '{dep_day}') {{
                        td.click();
                        return;
                    }}
                }}
            }})();
        """)
        await ke_page.wait_for_timeout(800)

        await ke_page.locator("input[autocomplete='family-name']").first.fill(last)
        await ke_page.wait_for_timeout(300)
        await ke_page.locator("input[autocomplete='given-name']").first.fill(first)
        await ke_page.wait_for_timeout(300)

        await ke_page.click("button:has-text('조회')", timeout=5000)

        try:
            await ke_page.wait_for_selector(".journey-info__date", timeout=20000)
        except:
            pass
        await ke_page.wait_for_timeout(2000)

        html_content = await ke_page.inner_text("body")

        if any(kw in html_content for kw in ["조회 결과가 없", "예약을 찾을 수 없", "확인되지 않", "일치하는 예약"]):
            return "❌ PNR오류", "예약 확인 불가"

        # 조회한 PNR이 실제로 결과 화면에 있는지 확인 (잔류 화면 오판정 방지 2중 안전장치)
        if pnr.upper() not in html_content.upper():
            return "❌ PNR오류", "조회한 PNR이 결과 화면에서 확인되지 않음"

        flt_match = re.search(r'KE\s*\d{3,4}', html_content)
        flt_found = flt_match.group().replace(" ", "") if flt_match else "편명미확인"

        date_match = re.search(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', html_content)
        if date_match:
            y = date_match.group(1)
            m = date_match.group(2).zfill(2)
            d = date_match.group(3).zfill(2)
            date_found = f"{y}-{m}-{d}"
        else:
            date_found = "날짜미확인"

        airports = re.findall(
            r'\b(PUS|GMP|ICN|CJU|TAE|CJJ|HIN|RSU|KPO|MWX'
            r'|NRT|HND|KIX|NGO|FUK|CTS|OKA'
            r'|BKK|CNX|HKT|SGN|HAN|DAD|CXR|MNL|CEB|CRK'
            r'|TPE|HKG|MFM|SIN|KUL|PNH|REP|VTE|RGN|DPS|PQC'
            r'|PEK|PVG|CAN|TAO|SZX|WEH|YNJ|HRB|SHE)\b',
            html_content
        )
        if len(airports) >= 2:
            route_found = f"{airports[0]}→{airports[1]}"
        else:
            route_found = "구간미확인"

        detail = f"{flt_found} | {date_found} | {route_found}"

        mismatch = []
        if dep_date and date_found != "날짜미확인":
            try:
                site_date = datetime.strptime(date_found, "%Y-%m-%d")
                if dep_date.date() != site_date.date():
                    mismatch.append(
                        f"날짜불일치(PDC:{dep_date.strftime('%m/%d')} vs 사이트:{site_date.strftime('%m/%d')})"
                    )
            except:
                pass

        if mismatch:
            return "⚠️ 불일치", detail + " | " + " / ".join(mismatch)

        if not _is_reliable_result(flt_found, route_found):
            return "❌ PNR오류", f"예약 확인 불가 (편명/구간 모두 미확인) | {detail}"

        # 같은 PNR 동승자 검증용으로 결과 페이지 보관
        _remember_page(pnr, target["airline"], html_content)
        return "✅ 확인완료", detail

    except PWTimeout:
        return "⏱️ 타임아웃", "재시도 필요"
    except Exception as e:
        logging.error(f"대한항공 조회 실패 | PNR: {pnr} | 탑승객: {last}{first}", exc_info=True)
        return "💥 오류", "시스템 로그 확인 필요"
    finally:
        await ke_page.close()


async def check_lj(page, target):
    pnr      = target["pnr"]
    last     = target["last"]
    first    = target["first"]
    eng_name = target.get("eng_name", "")
    dep      = target["dep"]
    arr      = target["arr"]

    dep_date = parse_dep_date(target["dep_time"])
    if not dep_date:
        return "💥 오류", "출발일 파싱 실패"

    intl = is_international(dep, arr)

    if intl and eng_name:
        parts       = eng_name.split("/")
        input_last  = parts[0].strip() if len(parts) >= 1 else last
        input_first = parts[1].strip() if len(parts) >= 2 else first
    else:
        input_last  = last
        input_first = first

    # 진에어는 SPA 구조 — 이전 조회 결과가 페이지에 잔류하므로 매 조회마다 새 탭 사용
    lj_page = await page.context.new_page()
    try:
        await lj_page.goto(LJ_URL, wait_until="domcontentloaded", timeout=20000)
        await lj_page.wait_for_timeout(2000)

        await lj_page.click("text=예약조회", timeout=5000)
        await lj_page.wait_for_timeout(800)

        await lj_page.locator("input[placeholder*='6자리']").first.fill(pnr)
        await lj_page.wait_for_timeout(300)

        await lj_page.locator("#lastName_resv").fill(input_last)
        await lj_page.wait_for_timeout(300)

        await lj_page.locator("#firstName_resv").fill(input_first)
        await lj_page.wait_for_timeout(300)

        await lj_page.click("#departureDate_resv", timeout=5000)
        await lj_page.wait_for_timeout(2000)

        dep_str = dep_date.strftime("%Y.%m.%d")
        iframe_locator = lj_page.frame_locator("iframe[src*='basicCalendarLayer']")

        try:
            date_input = iframe_locator.locator(f"input[name='hiddenDate'][value^='{dep_str}']")
            await date_input.evaluate("el => el.parentElement.click()")
        except:
            pass

        await lj_page.wait_for_timeout(800)
        await lj_page.click("button[role='login-button']", timeout=5000)

        # 진에어 내부 예약번호 ≠ 조회 PNR — "여정 예약정보" 텍스트로 성공 판정
        try:
            await lj_page.wait_for_selector("text=여정 예약정보", timeout=15000)
        except:
            pass
        await lj_page.wait_for_timeout(2000)

        html_content = await lj_page.inner_text("body")

        # CF 재체크
        if any(kw in html_content for kw in CF_KEYWORDS):
            return "⏱️ 타임아웃", "클라우드플레어 차단 → 재실행 필요"

        # 성공 판정: "여정 예약정보" 존재 여부
        if "여정 예약정보" not in html_content:
            if any(kw in html_content for kw in ["조회 결과가 없", "예약 내역이 없", "확인되지 않"]):
                return "❌ PNR오류", "예약 확인 불가"
            return "❌ PNR오류", "예약 확인 불가 (PNR 미조회)"

        flt_match = re.search(r'LJ\d{3,4}', html_content)
        flt_found = flt_match.group() if flt_match else "편명미확인"

        # 페이지에 "예약일 YYYY.MM.DD(요일)"가 실제 출발일보다 먼저 나와서
        # 첫 매치만 쓰면 예약일을 출발일로 착각함 — 예약일 다음 매치를 사용
        date_matches = re.findall(r'(\d{4})\.(\d{2})\.(\d{2})\(', html_content)
        if len(date_matches) >= 2:
            y, m, d = date_matches[1]
            date_found = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        elif len(date_matches) == 1:
            y, m, d = date_matches[0]
            date_found = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        else:
            date_found = "날짜미확인"

        airports = re.findall(
            r'(?<![A-Z0-9])(PUS|GMP|ICN|CJU|TAE|CJJ|HIN|RSU|KPO|MWX'
            r'|CNX|BKK|HKT|NRT|HND|KIX|NGO|CTS|FUK|OKA'
            r'|DAD|SGN|HAN|CXR|PQC|CEB|KLO|TAG|MNL'
            r'|TPE|HKG|MFM|SIN|DPS|GUM)(?![A-Z0-9])',
            html_content
        )
        if len(airports) >= 2:
            route_found = f"{airports[0]}→{airports[1]}"
        else:
            route_found = "구간미확인"

        detail = f"{flt_found} | {date_found} | {route_found}"

        mismatch = []
        if dep_date and date_found != "날짜미확인":
            try:
                site_date = datetime.strptime(date_found, "%Y-%m-%d")
                if dep_date.date() != site_date.date():
                    mismatch.append(
                        f"날짜불일치(PDC:{dep_date.strftime('%m/%d')} vs 사이트:{site_date.strftime('%m/%d')})"
                    )
            except:
                pass

        if mismatch:
            return "⚠️ 불일치", detail + " | " + " / ".join(mismatch)

        if not _is_reliable_result(flt_found, route_found):
            return "❌ PNR오류", f"예약 확인 불가 (편명/구간 모두 미확인) | {detail}"

        # 같은 PNR 동승자 검증용으로 결과 페이지 보관
        _remember_page(pnr, target["airline"], html_content)
        return "✅ 확인완료", detail

    except PWTimeout:
        return "⏱️ 타임아웃", "재시도 필요"
    except Exception as e:
        logging.error(f"진에어 조회 실패 | PNR: {pnr} | 탑승객: {input_last}{input_first}", exc_info=True)
        return "💥 오류", "시스템 로그 확인 필요"
    finally:
        await lj_page.close()


async def check_we(page, target, we_email):
    pnr      = target["pnr"]
    kor_name = target["kor_name"]

    if not we_email:
        return "⚠️ 수동확인필요", "파라타항공-이메일 미입력"

    dep_date = parse_dep_date(target["dep_time"])
    # dep_date 파싱 실패해도 조회는 진행 (날짜 불일치 검사만 스킵)

    # alert 팝업 대비 (틀린 PNR 등) - 메시지 캡처 후 자동 닫기
    dialog_msgs = []
    async def _on_dialog(dialog):
        dialog_msgs.append(dialog.message)
        await dialog.accept()
    page.on("dialog", _on_dialog)

    try:
        await page.goto(WE_URL, wait_until="domcontentloaded", timeout=20000)
        await page.wait_for_timeout(1500)

        # 홈페이지 비회원 탭 클릭
        await page.click("a[href='#nonmember']", timeout=5000)
        await page.wait_for_timeout(800)

        # 이메일 + 예약번호 입력
        await page.fill("#userEmail", we_email)
        await page.wait_for_timeout(300)
        await page.fill("#reservationNum", pnr)
        await page.wait_for_timeout(300)

        # 예약조회 버튼 클릭
        await page.click("#nonMemberResvSearchBtn", timeout=5000)

        # 결과 페이지 대기
        try:
            await page.wait_for_url("**/viewReservationDetail.do**", timeout=15000)
        except:
            pass
        await page.wait_for_timeout(2000)

        current_url = page.url
        html_content = await page.inner_text("body")

        # 실패 판정
        if "viewReservationDetail.do" not in current_url:
            if dialog_msgs:
                return "❌ PNR오류", f"알림: {dialog_msgs[0][:40]}"
            if any(kw in html_content for kw in ["일치하는 예약", "확인되지 않", "조회 결과가 없", "예약 내역이 없"]):
                return "❌ PNR오류", "예약 확인 불가"
            return "💥 오류", "결과 페이지 이동 실패"

        # 조회한 PNR이 실제로 결과 화면에 있는지 확인 (잔류 화면 오판정 방지 2중 안전장치)
        if pnr.upper() not in html_content.upper():
            return "❌ PNR오류", "조회한 PNR이 결과 화면에서 확인되지 않음"

        # 편명 파싱 (WE208 형태)
        flt_match = re.search(r'WE\s*\d{3,4}', html_content)
        flt_found = flt_match.group().replace(" ", "") if flt_match else "편명미확인"

        # 날짜 파싱: "출발일시" 라벨 뒤의 날짜만 (예약일 오인식 방지)
        date_match = re.search(r'출발일시[\s\S]{0,20}?(\d{4})\.(\d{2})\.(\d{2})', html_content)
        if not date_match:
            # fallback: 요일괄호가 바로 붙은 날짜 (출발일시 형식: 2026.06.15(월))
            date_match = re.search(r'(\d{4})\.(\d{2})\.(\d{2})\(', html_content)
        if date_match:
            date_found = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
        else:
            date_found = "날짜미확인"

        # 구간 파싱 (한글이 코드에 바로 붙어도 매칭: 푸꾸옥PQC서울 등)
        airports = re.findall(
            r'(?<![A-Z0-9])(PUS|GMP|ICN|CJU|TAE|CJJ|HIN|RSU|KPO|MWX|PQC|RGN|DPS|MFM|HKG)(?![A-Z0-9])',
            html_content
        )
        route_found = f"{airports[0]}→{airports[1]}" if len(airports) >= 2 else "구간미확인"

        detail = f"{flt_found} | {date_found} | {route_found}"

        # 날짜 불일치 검사
        mismatch = []
        if dep_date and date_found != "날짜미확인":
            try:
                site_date = datetime.strptime(date_found, "%Y-%m-%d")
                if dep_date.date() != site_date.date():
                    mismatch.append(
                        f"날짜불일치(PDC:{dep_date.strftime('%m/%d')} vs 사이트:{site_date.strftime('%m/%d')})"
                    )
            except:
                pass

        if mismatch:
            return "⚠️ 불일치", detail + " | " + " / ".join(mismatch)

        if not _is_reliable_result(flt_found, route_found):
            return "❌ PNR오류", f"예약 확인 불가 (편명/구간 모두 미확인) | {detail}"

        # 같은 PNR 동승자 검증용으로 결과 페이지 보관
        _remember_page(pnr, target["airline"], html_content)
        return "✅ 확인완료", detail

    except PWTimeout:
        return "⏱️ 타임아웃", "재시도 필요"
    except Exception:
        logging.error(f"파라타항공 조회 실패 | PNR: {pnr} | 탑승객: {kor_name}", exc_info=True)
        return "💥 오류", "시스템 로그 확인 필요"
    finally:
        page.remove_listener("dialog", _on_dialog)


async def check_tw(page, target):
    """
    티웨이항공 — Akamai 봇 차단으로 Playwright 자동 조회/자동입력 불가.
    Chrome 탭을 열고, 팝업에서 항목별 [복사] 버튼으로 사람이 직접 붙여넣도록 함.
    (자동 붙여넣기는 Akamai에 감지되어 사용 불가)
    """
    pnr      = target["pnr"]
    kor_name = target["kor_name"]
    eng_name = target.get("eng_name", "")
    dep_time = target.get("dep_time", "")
    dep      = target.get("dep", "")
    arr      = target.get("arr", "")

    # 자동화 창(Playwright)은 티웨이 작업에 쓰지 않음 — 혼동 방지용 안내 문구 표시
    try:
        await page.goto(
            "data:text/html,"
            "<html><body style='font-family:sans-serif;padding:60px;"
            "font-size:22px;color:#333;text-align:center;'>"
            "이 창은 자동화 전용입니다.<br><br>"
            "티웨이항공 조회는<br>"
            "<b>새로 열린 별도의 브라우저 창</b>에서 진행해주세요."
            "</body></html>",
            timeout=5000
        )
    except Exception:
        pass  # 안내 문구 표시 실패해도 조회 자체엔 영향 없음

    # 조회 페이지를 시스템 기본 브라우저의 새 창으로 오픈
    webbrowser.open(TW_URL, new=2)

    result_box = [None]
    popup = tk.Toplevel()
    popup.title("티웨이항공 수동 확인")
    popup.resizable(False, False)
    popup.attributes("-topmost", True)
    popup.grab_set()

    tk.Label(
        popup,
        text=(
            "티웨이항공은 보안 정책상 자동 조회가 불가합니다.\n"
            "방금 새로 열린 별도의 브라우저 창에서\n"
            "(자동화 창 아님 — about:blank 창은 무시하세요)\n"
            "아래 항목을 [복사] 버튼으로 복사해 붙여넣어\n"
            "직접 조회한 뒤 결과를 선택해주세요."
        ),
        justify="left", padx=20
    ).pack(pady=(15, 8))

    field_frame = tk.Frame(popup)
    field_frame.pack(padx=20, pady=5)

    def _copy(value):
        popup.clipboard_clear()
        popup.clipboard_append(value)

    def _add_field(row, label, value):
        tk.Label(field_frame, text=label, width=8, anchor="w").grid(row=row, column=0, sticky="w", pady=3)
        e = tk.Entry(field_frame, width=26)
        e.insert(0, value)
        e.config(state="readonly")
        e.grid(row=row, column=1, padx=6)
        tk.Button(field_frame, text="복사", width=6,
                  command=lambda v=value: _copy(v)).grid(row=row, column=2)

    fields = [("PNR", pnr), ("한글성명", kor_name)]
    if eng_name:
        parts = eng_name.split("/")
        eng_last  = parts[0].strip() if len(parts) >= 1 else eng_name
        eng_first = parts[1].strip() if len(parts) >= 2 else ""
        fields.append(("영문성", eng_last))
        if eng_first:
            fields.append(("영문이름", eng_first))
    fields.append(("구간", f"{dep} → {arr}"))
    fields.append(("출발일", dep_time))

    for i, (label, value) in enumerate(fields):
        _add_field(i, label, value)

    btn_frame = tk.Frame(popup)
    btn_frame.pack(pady=(10, 15))

    def _choose(v):
        result_box[0] = v
        popup.destroy()

    tk.Button(btn_frame, text="✅ 확인완료", width=12,
              command=lambda: _choose("ok")).pack(side="left", padx=6)
    tk.Button(btn_frame, text="❌ PNR오류", width=12,
              command=lambda: _choose("pnr_error")).pack(side="left", padx=6)
    tk.Button(btn_frame, text="⏭ 보류(건너뛰기)", width=14,
              command=lambda: _choose("skip")).pack(side="left", padx=6)

    popup.wait_window()

    if result_box[0] == "ok":
        return "✅ 확인완료", "[수동확인] 예약 확인됨"
    elif result_box[0] == "pnr_error":
        return "❌ PNR오류", "[수동확인] 예약 확인 불가"
    else:
        return "⚠️ 수동확인필요", "[수동확인] 보류됨 — 재확인 필요"


async def _dismiss_ad_popup(p):
    """
    그루비 광고 팝업 닫기 — 확인된 실물 셀렉터 하나만 정확히 겨냥.
    실물 img 태그: <img src="...groobee.io/image/close/..." alt="닫기" class="img_999999">
    실패해도 조회에 영향 없게 조용히 넘어감.
    """
    close_selector = "img[src*='groobee.io/image/close']"

    async def _try_close(scope):
        try:
            loc = scope.locator(close_selector).first
            if not await loc.is_visible(timeout=500):
                return False
            # 1) 일반 클릭 시도
            try:
                await loc.click(timeout=800)
                return True
            except Exception:
                pass
            # 2) 부모 요소 클릭 (img가 클릭 이벤트를 부모에 위임하는 경우)
            try:
                await loc.evaluate("el => (el.closest('a,button,div[onclick]') || el.parentElement).click()")
                return True
            except Exception:
                pass
            # 3) JS 직접 클릭 (오버레이/z-index 무시)
            try:
                await loc.evaluate("el => el.click()")
                return True
            except Exception:
                pass
        except Exception:
            pass
        return False

    # 최대 3회 시도 (팝업이 여러 개 겹쳐 있는 경우 대비)
    for _ in range(3):
        closed_any = False
        # 메인 프레임 + 모든 iframe 순회 (그루비는 iframe에 들어있기도 함)
        for scope in [p] + list(p.frames):
            if await _try_close(scope):
                closed_any = True
                await p.wait_for_timeout(400)
                break
        if not closed_any:
            break


def _prompt_jj_calendar(pnr, kor_name, target_date_display):
    """
    제주항공 달력 자동 클릭이 불안정해서, 입력값은 자동으로 채워두고
    달력 날짜 선택만 사람이 직접 하도록 안내하는 팝업.
    나머지(조회, 결과판정)는 계속 자동 진행됨.
    """
    result_box = [None]
    popup = tk.Toplevel()
    popup.title("제주항공 - 날짜 선택 필요")
    popup.resizable(False, False)
    popup.attributes("-topmost", True)
    popup.grab_set()

    tk.Label(
        popup,
        text=(
            "※ 제주항공은 보안 정책상 날짜 자동 선택이 안 됩니다.\n"
            "   (오류 아님 — 아래 순서대로만 해주시면 됩니다)\n\n"
            f"  탑승객 : {kor_name}\n"
            f"  PNR    : {pnr}\n"
            f"  출발일 : {target_date_display}\n\n"
            "① 지금 열려있는 자동화 Chrome 창을 클릭해서 앞으로 가져오세요\n"
            "   (이미 달력이 떠 있는 상태입니다)\n"
            "② 달력에서 위 '출발일' 날짜만 클릭하세요\n"
            "③ 아래 [완료] 버튼을 눌러주세요\n"
            "   → '선택'/'조회' 버튼 클릭 및 나머지는 자동으로 진행됩니다\n\n"
            "(혹시 실수로 선택/조회까지 눌러도 문제없이 진행됩니다)\n"
            "날짜가 헷갈리거나 실수했으면 [건너뛰기]를 눌러주세요."
        ),
        justify="left", padx=20, pady=15
    ).pack()

    btn_frame = tk.Frame(popup)
    btn_frame.pack(pady=(0, 15))

    def _choose(v):
        result_box[0] = v
        popup.destroy()

    tk.Button(btn_frame, text="✅ 완료 — 계속 진행", width=16,
              command=lambda: _choose("done")).pack(side="left", padx=6)
    tk.Button(btn_frame, text="⏭ 건너뛰기", width=12,
              command=lambda: _choose("skip")).pack(side="left", padx=6)

    popup.wait_window()
    return result_box[0]


async def check_jj(page, target):
    """
    제주항공 — 조회 페이지(viewOnOffReservationList.do)에서 입력 후
    결과 페이지(viewReservationDetail.do)로 이동.
    ※ 진입 시 마케팅 팝업이 뜰 수 있어 자동 닫기 시도 후 진행.
    ※ 달력 날짜 선택은 "일(day) 숫자" 텍스트 매칭 방식 — 검증 필요.
    """
    pnr      = target["pnr"]
    last     = target["last"]
    first    = target["first"]
    eng_name = target.get("eng_name", "")
    dep      = target["dep"]
    arr      = target["arr"]

    dep_date = parse_dep_date(target["dep_time"])
    if not dep_date:
        return "💥 오류", "출발일 파싱 실패"

    intl = is_international(dep, arr)

    if intl and eng_name:
        parts       = eng_name.split("/")
        input_last  = parts[0].strip() if len(parts) >= 1 else last
        input_first = parts[1].strip() if len(parts) >= 2 else first
    else:
        input_last  = last
        input_first = first

    jj_page = await page.context.new_page()
    try:
        await jj_page.goto(JJ_URL, wait_until="domcontentloaded", timeout=20000)
        await jj_page.wait_for_timeout(1500)

        # 진입 시 뜨는 마케팅 팝업(그루비 등)이 폼을 가려 클릭이 막히는 문제 방지
        await _dismiss_ad_popup(jj_page)

        await jj_page.locator("#recordLocatorLabel").fill(pnr)
        await jj_page.wait_for_timeout(300)

        await jj_page.locator("#psInputLastName_1").fill(input_last)
        await jj_page.wait_for_timeout(300)

        await jj_page.locator("#psInputFirstName_1").fill(input_first)
        await jj_page.locator("#psInputFirstName_1").press("Tab")
        await jj_page.wait_for_timeout(300)

        # 탑승일자 달력 선택 — 자동 클릭이 계속 불안정해서 사람이 직접 선택하도록 전환
        # (입력값은 이미 자동으로 채워짐, 달력 클릭만 사람이 하고 나머지는 자동 진행)
        await jj_page.click("#boardingDateBtn", timeout=5000)
        await jj_page.wait_for_selector("#datepicker01", timeout=5000)
        await jj_page.wait_for_timeout(500)
        await _dismiss_ad_popup(jj_page)

        target_date_display = dep_date.strftime("%Y-%m-%d (%a)")
        choice = _prompt_jj_calendar(pnr, target["kor_name"], target_date_display)

        if choice != "done":
            return "⚠️ 수동확인필요", "[제주달력] 사용자가 건너뜀"

        # 광고가 다시 떴을 수 있어 한 번 더 정리
        await _dismiss_ad_popup(jj_page)

        # 사람이 사이트 "선택" 버튼을 안 눌렀을 경우 대비해서 자동화가 눌러줌 (있으면)
        try:
            choose_btn = jj_page.locator("#chooseDepDateBtn")
            if await choose_btn.is_visible(timeout=500):
                await choose_btn.click(timeout=2000)
                await jj_page.wait_for_timeout(500)
        except Exception:
            pass

        # 상태 확인 — 두 가지 경우 모두 정상 진행:
        # (1) 검색 폼에서 날짜가 채워짐  (2) 이미 조회까지 눌러서 결과 페이지로 넘어감
        date_selected = False
        already_searched = False
        try:
            selectdate_val = await jj_page.evaluate(
                "() => { const el = document.querySelector('#selectDate'); return el ? el.value : null; }"
            )
            if selectdate_val:
                date_selected = True
            elif selectdate_val is None:
                # #selectDate 자체가 없다 = 검색 폼을 벗어난 것 (이미 결과 페이지로 이동)
                already_searched = True
        except Exception:
            pass

        if not date_selected and not already_searched:
            return "❌ PNR오류", "달력 날짜 선택 실패"

        await jj_page.wait_for_timeout(300)

        # 조회 버튼 — 사람이 이미 눌렀으면 중복 클릭하지 않음
        if not already_searched:
            # id="searchResvBtn" (입력 검증 통과 전까지 disabled)
            try:
                await jj_page.wait_for_selector("#searchResvBtn:not([disabled])", timeout=5000)
            except Exception:
                pass
            await jj_page.click("#searchResvBtn", timeout=5000)

        # 조회 성공 시 viewReservationDetail.do로 실제 페이지 이동(navigate)이 발생함
        # 텍스트만 기다리면 이동 중인 중간 상태를 잘못 캡처할 수 있어 URL 이동을 우선 대기
        # (URL이 안 바뀌는 SPA 방식일 수도 있어 짧게 시도 후 텍스트 대기로 넘어감)
        try:
            await jj_page.wait_for_url(re.compile(r"viewReservationDetail"), timeout=8000)
        except Exception:
            pass

        # URL 이동 후에도 렌더링 시간 필요 — "탑승객 정보" 텍스트로 최종 확인
        try:
            await jj_page.wait_for_selector("text=탑승객 정보", timeout=10000)
        except Exception:
            pass
        await jj_page.wait_for_timeout(1500)

        html_content = await jj_page.inner_text("body")

        if any(kw in html_content for kw in CF_KEYWORDS):
            return "⏱️ 타임아웃", "보안 확인 필요 → 재실행 필요"

        # 성공 판정: "탑승객 정보" 존재 여부
        if "탑승객 정보" not in html_content:
            return "❌ PNR오류", "예약 확인 불가"

        flt_match = re.search(r'7C\d{3,4}', html_content)
        flt_found = flt_match.group() if flt_match else "편명미확인"

        # 페이지에 "예약일 YYYY.MM.DD(요일)"가 실제 출발일보다 먼저 나와서
        # 첫 매치만 쓰면 예약일을 출발일로 착각함 — 예약일 다음 매치를 사용
        date_matches = re.findall(r'(\d{4})\.(\d{2})\.(\d{2})\(', html_content)
        if len(date_matches) >= 2:
            y, m, d = date_matches[1]
            date_found = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        elif len(date_matches) == 1:
            y, m, d = date_matches[0]
            date_found = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        else:
            date_found = "날짜미확인"

        # 이 사이트는 공항코드(PQC/ICN) 대신 한글 도시명("푸꾸옥","서울")으로 표시함
        try:
            city_titles = await jj_page.locator(".boarding__info-title").all_inner_texts()
            city_titles = [c.strip() for c in city_titles if c.strip()]
        except Exception:
            city_titles = []

        if len(city_titles) >= 2:
            route_found = f"{city_titles[0]}→{city_titles[1]}"
        else:
            # 혹시 코드로 표시되는 경우 대비한 폴백
            airports = re.findall(
                rf'(?<![A-Z0-9])({AIRPORT_CODES})(?![A-Z0-9])',
                html_content
            )
            if len(airports) >= 2:
                route_found = f"{airports[0]}→{airports[1]}"
            else:
                route_found = "구간미확인"

        detail = f"{flt_found} | {date_found} | {route_found}"

        mismatch = []
        if dep_date and date_found != "날짜미확인":
            try:
                site_date = datetime.strptime(date_found, "%Y-%m-%d")
                if dep_date.date() != site_date.date():
                    mismatch.append(
                        f"날짜불일치(PDC:{dep_date.strftime('%m/%d')} vs 사이트:{site_date.strftime('%m/%d')})"
                    )
            except Exception:
                pass

        if mismatch:
            return "⚠️ 불일치", detail + " | " + " / ".join(mismatch)

        if not _is_reliable_result(flt_found, route_found):
            return "❌ PNR오류", f"예약 확인 불가 (편명/구간 모두 미확인) | {detail}"

        # 같은 PNR 동승자 검증용으로 결과 페이지 보관
        _remember_page(pnr, target["airline"], html_content)
        return "✅ 확인완료", detail

    except PWTimeout:
        return "⏱️ 타임아웃", "재시도 필요"
    except Exception:
        logging.error(f"제주항공 조회 실패 | PNR: {pnr} | 탑승객: {input_last}{input_first}", exc_info=True)
        return "💥 오류", "시스템 로그 확인 필요"
    finally:
        await jj_page.close()


async def run_check(page, target, we_email=""):
    """단일 조회 실행 + 재시도 로직"""
    airline  = target["airline"]
    eng_name = target.get("eng_name", "")

    if airline == "에어부산":
        result, detail = await check_bx(page, target)
    elif airline == "대한항공":
        result, detail = await check_ke(page, target)
    elif airline == "진에어":
        result, detail = await check_lj(page, target)
    elif airline == "제주항공":
        result, detail = await check_jj(page, target)
    elif airline == "파라타항공":
        result, detail = await check_we(page, target, we_email)
    elif airline == "티웨이항공":
        # 수동확인 방식 — 자동 재시도/영문재시도 로직 대상 아님
        return await check_tw(page, target)
    else:
        return "⬜ 미지원", "지원 항공사 아님"

    # ── 국내선 한글 조회 실패 시 영문으로 재시도 (외국인 승무원 대비) ──
    # 파라타 제외 / 영문명 있을 때 / PNR오류·예약없음일 때만
    intl = is_international(target["dep"], target["arr"])
    if (
        airline in ("에어부산", "대한항공", "진에어", "제주항공")
        and not intl
        and eng_name
        and any(kw in result for kw in ["PNR오류", "예약없음"])
    ):
        # 영문 이름으로 강제 전환한 임시 target 복사
        parts = eng_name.split("/")
        tmp = dict(target)
        tmp["last"]  = parts[0].strip() if len(parts) >= 1 else target["last"]
        tmp["first"] = parts[1].strip() if len(parts) >= 2 else target["first"]
        # 국제선 분기를 타지 않도록 dep/arr을 해외로 속이지 않고
        # check_* 함수 내 intl=False → 한글 사용이 되므로
        # last/first만 덮어쓴 tmp를 넘기면 영문으로 입력됨
        await asyncio.sleep(1)
        r2, d2 = result, detail   # 분기 누락 시 NameError 방지
        if airline == "에어부산":
            r2, d2 = await check_bx(page, tmp)
        elif airline == "대한항공":
            r2, d2 = await check_ke(page, tmp)
        elif airline == "진에어":
            r2, d2 = await check_lj(page, tmp)
        elif airline == "제주항공":
            r2, d2 = await check_jj(page, tmp)
        if "확인완료" in r2 or "불일치" in r2:
            result = r2
            detail = "[영문재시도] " + d2

    # ── 타임아웃/오류 시 1회 재시도 ──
    elif "타임아웃" in result or ("오류" in result and "PNR" not in result and "파싱" not in detail):
        await asyncio.sleep(2)
        if airline == "에어부산":
            result, detail = await check_bx(page, target)
        elif airline == "대한항공":
            result, detail = await check_ke(page, target)
        elif airline == "진에어":
            result, detail = await check_lj(page, target)
        elif airline == "제주항공":
            result, detail = await check_jj(page, target)
        elif airline == "파라타항공":
            result, detail = await check_we(page, target, we_email)
        if "확인완료" in result or "불일치" in result:
            detail = "[재시도 성공] " + detail

    return result, detail


async def main():
    os.system(f"title 타사 예약 자동 검증 시스템 v{__version__}")
    print(f"{'='*50}")
    print(f"✈️  타사 예약 자동 검증 시스템 v{__version__}")
    print("문의: 승무계획팀")
    print(f"{'='*50}")
    print(f"오류 로그 저장 위치: {LOG_PATH}")
    print(f"\n[이번 버전 변경사항]\n{LATEST_CHANGELOG}\n")
    print(f"{'='*50}\n")

    # 조회 범위 선택 팝업
    mode, start_date, end_date, delay_min, delay_max = get_check_mode()
    if mode == "5d":
        mode_label = "오늘~5일 이내"
    elif mode == "this":
        mode_label = f"이번달 말까지 ({end_date.month}월 {end_date.day}일)"
    else:
        mode_label = f"다음달 ({start_date.month}월 {start_date.day}일 ~ {end_date.month}월 {end_date.day}일)"

    targets = load_targets(EXCEL_PATH, SHEET_NAME, start_date, end_date)
    total   = len(targets)

    print(f"📂 선택한 파일: {EXCEL_PATH}")
    print(f"   시트: {SHEET_NAME}\n")

    if total == 0:
        print(f"검증 대상이 없습니다. ({mode_label})")
        print("\n⚠️  데이터가 분명히 있는데 0건이면 아래를 확인하세요:")
        print("   1) 위에 표시된 '선택한 파일'이 실제로 데이터를 넣은 그 파일이 맞나요?")
        print("      (같은 이름의 다른 파일/예전 파일을 골랐을 수 있습니다)")
        print("   2) 엑셀에서 데이터 입력 후 저장(Ctrl+S)했나요?")
        print(f"   3) 조회 범위({mode_label})에 실제 출발일이 포함되나요?")
        input("\n엔터 누르면 종료...")
        return

    # ── 이미 조회 완료된 건이 있으면 이어서 할지 확인 ──
    done_targets = [t for t in targets if "확인완료" in t["prev_result"]]
    if done_targets:
        if ask_resume(len(done_targets), total):
            for t in done_targets:
                t["result"] = t["prev_result"]
                t["detail"] = t["prev_detail"]
                t["skipped"] = True
            print(f"⏭  이미 완료된 {len(done_targets)}건은 건너뜁니다. "
                  f"(남은 조회: {total - len(done_targets)}건)\n")

    pending = [t for t in targets if not t.get("skipped")]

    bx_cnt = sum(1 for t in pending if t["airline"] == "에어부산")
    ke_cnt = sum(1 for t in pending if t["airline"] == "대한항공")
    lj_cnt = sum(1 for t in pending if t["airline"] == "진에어")
    jj_cnt = sum(1 for t in pending if t["airline"] == "제주항공")
    we_cnt = sum(1 for t in pending if t["airline"] == "파라타항공")
    tw_cnt = sum(1 for t in pending if t["airline"] == "티웨이항공")

    print(f"검증 대상: {len(pending)}건 ({mode_label})")
    print(f"  에어부산: {bx_cnt}건 | 대한항공: {ke_cnt}건 | 진에어: {lj_cnt}건 | 제주항공: {jj_cnt}건 | 파라타항공: {we_cnt}건 | 티웨이: {tw_cnt}건")
    print(f"  딜레이: {delay_min}~{delay_max}초")
    print(f"{'='*50}\n")

    # 건너뛴 것만 있고 조회할 게 없으면 바로 저장하고 종료
    if not pending:
        print("모두 이미 조회 완료된 상태입니다. 결과만 다시 저장합니다.\n")
        save_results(EXCEL_PATH, SHEET_NAME, targets)
        input("\n엔터 누르면 종료...")
        return

    # 파라타항공 건수 있으면 이메일 입력 팝업
    we_email = ""
    if we_cnt > 0:
        from tkinter import simpledialog
        we_email = simpledialog.askstring(
            "파라타항공 조회 이메일",
            f"파라타항공 예약 {we_cnt}건이 있습니다.\n\n예약 시 사용한 이메일 주소를 입력하세요.\n(홈페이지 비회원 조회용)",
            parent=root
        )
        if not we_email:
            print("⚠️  이메일 미입력 → 파라타항공 건은 수동확인으로 처리됩니다.")
        else:
            print(f"파라타항공 조회 이메일: {we_email}\n")

    # 시스템 Chrome 경로 자동 탐색
    CHROME_PATHS = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe"),
    ]
    chrome_exe = next((p for p in CHROME_PATHS if os.path.exists(p)), None)

    # Chrome 프로필 경로 (쿠키·히스토리 재사용 → 클라우드플레어 신뢰 점수 향상)
    chrome_profile = os.path.expanduser(r"~\AppData\Local\Google\Chrome\User Data")
    use_profile = chrome_exe and os.path.exists(chrome_profile)

    if chrome_exe:
        print(f"시스템 Chrome 사용: {chrome_exe}")
    else:
        print("⚠️  Chrome 미발견 → Playwright Chromium으로 실행 (에어부산 캡챠 발생 가능)")

    # Chrome 쿠키 파일 임시 복사 (원본 잠금 회피)
    import shutil, tempfile
    async with async_playwright() as p:
        launch_kwargs = dict(
            headless=HEADLESS,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-infobars",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-extensions",
                "--window-size=1280,800",
            ]
        )
        if chrome_exe:
            launch_kwargs["executable_path"] = chrome_exe

        browser = await p.chromium.launch(**launch_kwargs)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="ko-KR",
            timezone_id="Asia/Seoul",
            viewport={"width": 1280, "height": 800},
            java_script_enabled=True,
        )

        # playwright-stealth 적용 (클라우드플레어 핑거프린트 우회)
        if STEALTH_AVAILABLE:
            stealth = Stealth(
                navigator_languages_override=("ko-KR", "ko"),
                navigator_platform_override="Win32",
                navigator_webdriver=True,
                chrome_runtime=True,
            )
            print("playwright-stealth 적용 완료 (클라우드플레어 우회 시도)")
        else:
            stealth = None
            await context.add_init_script("""
                Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
                Object.defineProperty(navigator, 'languages', {get: () => ['ko-KR','ko','en-US','en']});
                window.chrome = {runtime: {}};
            """)
            print("⚠️  playwright-stealth 미설치 → 기본 우회 모드")

        page = await context.new_page()

        # stealth를 page에 적용
        if stealth:
            await stealth.apply_stealth_async(page)

        pnr_cache = {}

        pending_total = len(pending)
        for i, target in enumerate(pending, 1):
            airline = target["airline"]
            pnr     = target["pnr"]
            print(f"[{i:02d}/{pending_total}] {target['kor_name']:5} | {airline} | {pnr} | ", end="", flush=True)

            cache_key = (pnr, airline)
            # 캐시는 같은 PNR일 때만 잡히지만, 그 예약에 이 사람이 실제로
            # 포함돼 있는지까지 확인해야 함 (이름 확인 없이 통과하면 오판정)
            use_cache = False
            if cache_key in pnr_cache:
                if _name_in_page(target, _page_cache.get(cache_key, "")):
                    use_cache = True

            if use_cache:
                result, detail = pnr_cache[cache_key]
                target["result"] = result
                target["detail"] = detail
                print(f"[캐시·동승확인] {result}  {detail}")
            else:
                # 예상 못한 오류(브라우저 강제종료 등)로 전체가 중단되지 않도록 보호
                # → 여기까지 조회한 결과는 반드시 엑셀에 저장됨
                try:
                    result, detail = await run_check(page, target, we_email)
                except Exception as exc:
                    logging.error(f"조회 중 예외 | {airline} | PNR: {pnr}", exc_info=True)
                    result, detail = "💥 오류", "조회 중 오류 발생 (로그 확인)"
                    if "closed" in str(exc).lower():
                        target["result"] = result
                        target["detail"] = detail
                        print(f"{result}  {detail}")
                        print("\n⚠️  브라우저가 닫혀 남은 건은 진행할 수 없습니다.")
                        print("→ 여기까지의 결과는 엑셀에 저장됩니다.\n")
                        for rest in pending[i:]:
                            rest["result"] = "⬜ 미조회"
                            rest["detail"] = "브라우저 종료로 미처리"
                        break
                target["result"] = result
                target["detail"] = detail
                # 성공한 결과만 캐시에 저장
                if "확인완료" in result:
                    pnr_cache[cache_key] = (result, detail)
                print(f"{result}  {detail}")
                if i < pending_total:
                    # 에어부산은 클라우드플레어 대비 딜레이 더 늘림
                    if target["airline"] == "에어부산":
                        await asyncio.sleep(random.uniform(5.0, 10.0))
                    else:
                        await asyncio.sleep(random.uniform(delay_min, delay_max))

        try:
            await browser.close()
        except Exception:
            pass  # 이미 닫혀있으면 무시

    save_results(EXCEL_PATH, SHEET_NAME, targets)

    confirmed = sum(1 for t in targets if t["result"] and "확인완료"    in str(t["result"]))
    mismatch  = sum(1 for t in targets if t["result"] and "불일치"       in str(t["result"]))
    no_rsv    = sum(1 for t in targets if t["result"] and "예약없음"     in str(t["result"]))
    pnr_err   = sum(1 for t in targets if t["result"] and "PNR오류"     in str(t["result"]))
    manual    = sum(1 for t in targets if t["result"] and "수동확인필요" in str(t["result"]))
    # "오류"는 "PNR오류"의 부분문자열이라 그대로 세면 중복 집계됨 → PNR오류는 제외
    error     = sum(1 for t in targets if t["result"] and
                    (("오류" in str(t["result"]) and "PNR오류" not in str(t["result"]))
                     or "타임아웃" in str(t["result"])))
    skipped   = sum(1 for t in targets if t["result"] and "미조회" in str(t["result"]))

    print(f"\n{'='*50}")
    print(f"✅ 확인완료      : {confirmed}건")
    print(f"⚠️  불일치       : {mismatch}건  ← 즉시 확인!")
    print(f"❌ 예약없음      : {no_rsv}건   ← 즉시 확인!")
    print(f"❌ PNR오류       : {pnr_err}건  ← 즉시 확인!")
    print(f"⚠️  수동확인필요  : {manual}건  ← 파라타항공 직접 조회 필요!")
    print(f"💥 오류/재시도   : {error}건")
    if skipped:
        print(f"⬜ 미조회        : {skipped}건  ← 중단되어 조회 못함, 재실행 필요!")
    input("\n엔터 누르면 종료...")


if __name__ == "__main__":
    asyncio.run(main())